#!/usr/bin/env python3
"""엑셀 → 마크다운 변환기 (이미지 추출 지원)

사용법:
    python excel2md.py report.xlsx
    python excel2md.py data.xlsx --sheet "매출현황"
    python excel2md.py data.xlsx --output ./docs --overwrite
    python excel2md.py data.xlsx --no-images
"""

import argparse
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook


# xlsx 내부 XML 네임스페이스
NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
}


def extract_images(xlsx_path: Path, out_dir: Path) -> dict:
    """xlsx에서 이미지를 추출하고 시트별/행별로 매핑

    Args:
        xlsx_path: xlsx 파일 경로
        out_dir: 이미지 저장 디렉토리

    Returns:
        {sheet_index(0-based): [{'filename': str, 'row': int, 'col': int}]}
        행 매핑 실패 시 row=-1로 표시
    """
    result = defaultdict(list)

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            zip_names = zf.namelist()

            # 1. 시트→드로잉 매핑 수집
            # xl/worksheets/_rels/sheet1.xml.rels → drawing1.xml
            sheet_to_drawing = {}
            sheet_rels_pattern = re.compile(r'xl/worksheets/_rels/sheet(\d+)\.xml\.rels')

            for name in zip_names:
                m = sheet_rels_pattern.match(name)
                if not m:
                    continue
                sheet_idx = int(m.group(1)) - 1  # 0-based
                rels_xml = zf.read(name).decode('utf-8')
                rels_root = ET.fromstring(rels_xml)
                for rel in rels_root.findall('rel:Relationship', NS):
                    rel_type = rel.get('Type', '')
                    if 'drawing' in rel_type:
                        target = rel.get('Target', '')
                        # ../drawings/drawing1.xml → drawing1.xml
                        drawing_name = target.split('/')[-1]
                        sheet_to_drawing[sheet_idx] = drawing_name

            if not sheet_to_drawing:
                # 드로잉이 없으면 fallback: media 폴더만 추출
                return _extract_media_fallback(zf, out_dir)

            # 2. 각 드로잉 파일 처리
            for sheet_idx, drawing_name in sheet_to_drawing.items():
                drawing_path = f'xl/drawings/{drawing_name}'
                rels_path = f'xl/drawings/_rels/{drawing_name}.rels'

                if drawing_path not in zip_names:
                    continue

                # 2a. rId → 미디어 파일 매핑
                rid_to_media = {}
                if rels_path in zip_names:
                    rels_xml = zf.read(rels_path).decode('utf-8')
                    rels_root = ET.fromstring(rels_xml)
                    for rel in rels_root.findall('rel:Relationship', NS):
                        rel_type = rel.get('Type', '')
                        if 'image' in rel_type:
                            rid = rel.get('Id', '')
                            target = rel.get('Target', '')
                            # ../media/image1.png → xl/media/image1.png
                            media_path = 'xl/media/' + target.split('/')[-1]
                            rid_to_media[rid] = media_path

                # 2b. 드로잉 XML에서 앵커 파싱 (row, col, rId)
                drawing_xml = zf.read(drawing_path).decode('utf-8')
                drawing_root = ET.fromstring(drawing_xml)

                # twoCellAnchor와 oneCellAnchor 모두 처리
                anchors = (
                    drawing_root.findall('xdr:twoCellAnchor', NS) +
                    drawing_root.findall('xdr:oneCellAnchor', NS)
                )

                for anchor in anchors:
                    from_el = anchor.find('xdr:from', NS)
                    if from_el is None:
                        continue

                    row_el = from_el.find('xdr:row', NS)
                    col_el = from_el.find('xdr:col', NS)
                    if row_el is None or col_el is None:
                        continue

                    row = int(row_el.text)
                    col = int(col_el.text)

                    # blip에서 rId 추출
                    blip = anchor.find('.//a:blip', NS)
                    if blip is None:
                        continue

                    rid = blip.get(f'{{{NS["r"]}}}embed')
                    if not rid or rid not in rid_to_media:
                        continue

                    media_zip_path = rid_to_media[rid]
                    if media_zip_path not in zip_names:
                        continue

                    # 2c. 이미지 파일 추출
                    original_name = media_zip_path.split('/')[-1]
                    img_data = zf.read(media_zip_path)
                    img_path = out_dir / original_name
                    img_path.write_bytes(img_data)

                    result[sheet_idx].append({
                        'filename': original_name,
                        'row': row,
                        'col': col,
                    })
                    print(f"  [IMG] {original_name} → row {row}, col {col}")

            # 각 시트의 이미지를 row 순으로 정렬
            for sheet_idx in result:
                result[sheet_idx].sort(key=lambda x: (x['row'], x['col']))

    except zipfile.BadZipFile:
        print("[WARN] xlsx 파일을 ZIP으로 열 수 없습니다 (이미지 추출 건너뜀)")
    except Exception as e:
        print(f"[WARN] 이미지 추출 중 오류: {e} (이미지 추출 건너뜀)")

    return dict(result)


def _extract_media_fallback(zf: zipfile.ZipFile, out_dir: Path) -> dict:
    """드로잉 매핑 없이 media 폴더 전체 추출 (위치 정보 없음)"""
    result = defaultdict(list)
    for name in zf.namelist():
        if name.startswith('xl/media/'):
            original_name = name.split('/')[-1]
            img_data = zf.read(name)
            img_path = out_dir / original_name
            img_path.write_bytes(img_data)
            # sheet 0에 row=-1(매핑 불가)로 추가
            result[0].append({
                'filename': original_name,
                'row': -1,
                'col': -1,
            })
            print(f"  [IMG] {original_name} (위치 미확인)")
    return dict(result)


def excel_to_markdown(
    excel_path: str,
    output_dir: str | None = None,
    sheet_name: str | None = None,
    overwrite: bool = False,
    no_images: bool = False
) -> list[str]:
    """엑셀 파일을 마크다운으로 변환

    Args:
        excel_path: 엑셀 파일 경로
        output_dir: 출력 디렉토리 (기본: 엑셀파일명 폴더)
        sheet_name: 특정 시트만 변환 (None이면 전체)
        overwrite: 기존 파일 덮어쓰기
        no_images: True면 이미지 추출 건너뛰기

    Returns:
        생성된 md 파일 경로 목록
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_path}")

    # 출력 디렉토리 설정
    if output_dir:
        out_dir = Path(output_dir) / excel_path.stem
    else:
        out_dir = excel_path.parent / excel_path.stem

    # 디렉토리 생성
    if out_dir.exists() and not overwrite:
        print(f"[WARN] 폴더가 이미 존재합니다: {out_dir}")
        print("       --overwrite 옵션으로 덮어쓰기 가능")
        return []

    out_dir.mkdir(parents=True, exist_ok=True)

    # 이미지 추출
    image_map = {}
    if not no_images:
        image_map = extract_images(excel_path, out_dir)
        total_images = sum(len(v) for v in image_map.values())
        if total_images > 0:
            print(f"[OK] 이미지 {total_images}개 추출 완료")

    # 엑셀 로드 (data_only=True로 수식 결과값 읽기)
    wb = load_workbook(excel_path, data_only=True)

    created_files = []

    # 시트 필터링
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for idx, name in enumerate(wb.sheetnames):
        if name not in sheets:
            continue

        ws = wb[name]
        # 해당 시트의 이미지 목록
        sheet_images = image_map.get(idx, [])
        md_content = sheet_to_markdown(ws, excel_path.name, name, sheet_images)

        # 파일명에서 특수문자 제거
        safe_name = sanitize_filename(name)
        md_path = out_dir / f"{safe_name}.md"

        md_path.write_text(md_content, encoding='utf-8')
        created_files.append(str(md_path))
        img_note = f" (+이미지 {len(sheet_images)}개)" if sheet_images else ""
        print(f"[OK] {md_path}{img_note}")

    wb.close()
    return created_files


def sheet_to_markdown(ws, source_file: str, sheet_name: str,
                      images: list | None = None) -> str:
    """워크시트를 마크다운 문자열로 변환

    Args:
        ws: openpyxl 워크시트
        source_file: 원본 파일명
        sheet_name: 시트명
        images: 이미지 목록 [{'filename', 'row', 'col'}] (없으면 None/빈 리스트)
    """
    if images is None:
        images = []

    lines = []

    # 헤더
    lines.append(f"# {sheet_name}")
    lines.append("")

    # 메타 정보 (이미지 수 포함)
    row_count = ws.max_row
    col_count = ws.max_column
    meta = f"> Source: {source_file} | Sheet: {sheet_name} | Rows: {row_count}"
    if images:
        meta += f" | Images: {len(images)}"
    lines.append(meta)
    lines.append("")

    if row_count == 0 or col_count == 0:
        lines.append("_(빈 시트)_")
        return "\n".join(lines)

    # 데이터 추출
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        lines.append("_(데이터 없음)_")
        return "\n".join(lines)

    # 첫 행을 헤더로 사용
    headers = rows[0]
    data_rows = rows[1:]

    # 빈 헤더 처리
    headers = [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(headers)]

    # 마크다운 테이블 생성
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join(["---"] * len(headers)) + "|")

    for row in data_rows:
        # None을 빈 문자열로, 나머지는 문자열로 변환
        cells = [format_cell(cell) for cell in row]
        # 컬럼 수 맞추기
        while len(cells) < len(headers):
            cells.append("")
        lines.append("| " + " | ".join(cells[:len(headers)]) + " |")

    # 이미지 섹션 추가
    if images:
        lines.append("")
        lines.append("## 첨부 이미지")
        lines.append("")

        # 행별로 그룹핑 (row 0 = 헤더 → 데이터 행 번호로 변환)
        mapped_images = defaultdict(list)
        unmapped_images = []

        for img in images:
            if img['row'] < 0:
                unmapped_images.append(img)
            else:
                # xlsx row는 0-based, row 0 = 헤더, row 1 = 데이터 1행
                data_row = img['row']  # 0-based → 표시는 그대로 사용
                mapped_images[data_row].append(img)

        # 행 순서로 출력
        for data_row in sorted(mapped_images.keys()):
            imgs = mapped_images[data_row]
            lines.append(f"### 행 {data_row}")
            lines.append("")
            for img in imgs:
                lines.append(f"![{img['filename']}]({img['filename']})")
                lines.append("")

        # 매핑 불가 이미지
        if unmapped_images:
            lines.append("### 기타")
            lines.append("")
            for img in unmapped_images:
                lines.append(f"![{img['filename']}]({img['filename']})")
                lines.append("")

    return "\n".join(lines)


def format_cell(value) -> str:
    """셀 값을 마크다운에 적합한 문자열로 변환"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float):
        # 정수면 소수점 제거
        if value == int(value):
            return f"{int(value):,}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    # 파이프 문자 이스케이프 (마크다운 테이블 깨짐 방지)
    return str(value).replace("|", "\\|").replace("\n", " ")


def sanitize_filename(name: str) -> str:
    """파일명에 사용할 수 없는 문자 제거"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()


def main():
    parser = argparse.ArgumentParser(
        description="엑셀 파일을 마크다운 테이블로 변환 (이미지 추출 지원)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python excel2md.py report.xlsx
  python excel2md.py data.xlsx --sheet "매출현황"
  python excel2md.py data.xlsx --output ./docs --overwrite
  python excel2md.py data.xlsx --no-images
        """
    )

    parser.add_argument("excel_file", help="변환할 엑셀 파일 경로")
    parser.add_argument("--sheet", "-s", help="특정 시트만 변환")
    parser.add_argument("--output", "-o", help="출력 디렉토리")
    parser.add_argument("--overwrite", "-f", action="store_true",
                        help="기존 폴더 덮어쓰기")
    parser.add_argument("--no-images", action="store_true",
                        help="이미지 추출 건너뛰기")

    args = parser.parse_args()

    try:
        files = excel_to_markdown(
            args.excel_file,
            output_dir=args.output,
            sheet_name=args.sheet,
            overwrite=args.overwrite,
            no_images=args.no_images
        )

        if files:
            print(f"\n[DONE] {len(files)} files created")

    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
