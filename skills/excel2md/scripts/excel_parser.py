# requires: openpyxl
# pip install openpyxl

"""
openpyxl 기반 Excel 파싱 헬퍼 함수 라이브러리.

엑셀 파일을 읽고, 마크다운 테이블 / JSON 으로 변환하거나
임베디드 이미지를 추출하는 공통 함수를 제공합니다.
excel2md.py 의 핵심 로직을 독립 함수로 분리하여 재사용을 용이하게 합니다.
"""

import json
import os
import zipfile
import xml.etree.ElementTree as ET
from typing import Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


# ──────────────────────────────────────────────────────────────────────────────
# 워크북 열기
# ──────────────────────────────────────────────────────────────────────────────

def read_workbook(filepath: str, data_only: bool = True) -> Workbook:
    """
    openpyxl로 엑셀 워크북을 엽니다.

    수식 셀은 data_only=True(기본값)일 때 마지막으로 계산된 결과값만 반환됩니다.
    수식 원본을 보려면 data_only=False로 전달하세요.

    Args:
        filepath:  엑셀 파일 경로 (.xlsx)
        data_only: True → 수식 결과값 반환, False → 수식 원본 반환

    Returns:
        openpyxl Workbook 객체
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")
    return load_workbook(filepath, data_only=data_only)


# ──────────────────────────────────────────────────────────────────────────────
# 헤더 자동 감지
# ──────────────────────────────────────────────────────────────────────────────

def detect_header_row(sheet: Worksheet) -> int:
    """
    헤더 행 번호(1-based)를 자동으로 감지합니다.

    헤더 행 판별 기준:
    1. 비어있지 않은 첫 번째 행을 후보로 선택합니다.
    2. 해당 행에서 문자열 셀 비율이 50% 이상이면 헤더로 판정합니다.
    3. 판정 실패 시 1행을 반환합니다.

    Args:
        sheet: openpyxl Worksheet 객체

    Returns:
        헤더가 있는 행 번호 (1-based)
    """
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        non_empty = [c for c in row if c.value is not None]
        if not non_empty:
            continue  # 빈 행 건너뜀

        str_count = sum(1 for c in non_empty if isinstance(c.value, str))
        if len(non_empty) > 0 and str_count / len(non_empty) >= 0.5:
            return row_idx
        # 첫 비어있지 않은 행이 문자열 비율 미달이면 그냥 1행 반환
        return row_idx

    return 1


# ──────────────────────────────────────────────────────────────────────────────
# 시트 → 마크다운
# ──────────────────────────────────────────────────────────────────────────────

def sheet_to_markdown(sheet: Worksheet, header_row: int | None = None) -> str:
    """
    openpyxl Worksheet를 GitHub Flavored Markdown 테이블 문자열로 변환합니다.

    Args:
        sheet:      변환할 Worksheet
        header_row: 헤더 행 번호(1-based). None이면 detect_header_row()로 자동 감지.

    Returns:
        마크다운 테이블 문자열 (헤더 포함)
    """
    if header_row is None:
        header_row = detect_header_row(sheet)

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return ""

    # 헤더 행 추출
    headers = [str(v) if v is not None else "" for v in all_rows[header_row - 1]]
    col_count = len(headers)

    lines: list[str] = []
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * col_count) + " |")

    # 데이터 행
    for row in all_rows[header_row:]:
        # 빈 행 건너뜀
        if all(v is None for v in row):
            continue
        cells = []
        for v in row:
            cell_str = _format_cell(v)
            cells.append(cell_str.replace("|", "\\|"))  # 파이프 이스케이프
        # 열 수 맞춤
        while len(cells) < col_count:
            cells.append("")
        lines.append("| " + " | ".join(cells[:col_count]) + " |")

    return "\n".join(lines)


def _format_cell(value: Any) -> str:
    """셀 값을 마크다운용 문자열로 변환하는 내부 헬퍼."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float):
        # 정수처럼 보이는 소수 처리
        if value == int(value):
            return f"{int(value):,}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


# ──────────────────────────────────────────────────────────────────────────────
# 시트 → JSON
# ──────────────────────────────────────────────────────────────────────────────

def sheet_to_json(sheet: Worksheet, header_row: int | None = None) -> dict:
    """
    openpyxl Worksheet를 구조화된 dict로 변환합니다.

    반환 형식:
    {
        "sheet": "Sheet1",
        "row_count": 9,
        "headers": ["순번", "항목", "금액"],
        "rows": [
            {"순번": 1, "항목": "개발비", "금액": 5000000},
            ...
        ]
    }

    Args:
        sheet:      변환할 Worksheet
        header_row: 헤더 행 번호(1-based). None이면 자동 감지.

    Returns:
        구조화된 dict
    """
    if header_row is None:
        header_row = detect_header_row(sheet)

    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        return {"sheet": sheet.title, "row_count": 0, "headers": [], "rows": []}

    headers = [str(v) if v is not None else f"Col{i}" for i, v in enumerate(all_rows[header_row - 1])]
    col_count = len(headers)

    rows = []
    for row in all_rows[header_row:]:
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, v in enumerate(row[:col_count]):
            key = headers[i]
            row_dict[key] = v
        rows.append(row_dict)

    return {
        "sheet": sheet.title,
        "row_count": len(rows),
        "headers": headers,
        "rows": rows,
    }


# ──────────────────────────────────────────────────────────────────────────────
# 임베디드 이미지 추출
# ──────────────────────────────────────────────────────────────────────────────

def extract_images(filepath: str, output_dir: str) -> list[dict]:
    """
    xlsx 파일에서 임베디드 이미지를 추출하고 행/열 위치 정보를 반환합니다.

    xlsx는 ZIP 아카이브이므로, 내부 XML을 직접 파싱하여 이미지를 꺼냅니다.
    드로잉 매핑 파싱에 실패하면 media 전체를 추출하고 row=-1로 표시합니다.

    Args:
        filepath:   원본 xlsx 파일 경로
        output_dir: 이미지를 저장할 디렉토리 경로 (없으면 자동 생성)

    Returns:
        이미지 정보 목록.
        [{"filename": "image1.png", "row": 0, "col": 0}, ...]
        행/열을 알 수 없는 경우 row=-1, col=-1
    """
    os.makedirs(output_dir, exist_ok=True)
    results: list[dict] = []

    ns = {
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    }

    with zipfile.ZipFile(filepath, "r") as zf:
        names = zf.namelist()

        # ── 드로잉 기반 추출 ─────────────────────────────────
        drawing_files = [n for n in names if n.startswith("xl/drawings/drawing") and n.endswith(".xml")]
        drawing_rels: dict[str, dict[str, str]] = {}

        for drawing_path in drawing_files:
            rels_path = drawing_path.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
            if rels_path not in names:
                continue
            tree = ET.parse(zf.open(rels_path))
            root = tree.getroot()
            rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
            rid_map: dict[str, str] = {}
            for rel in root.findall(f"{{{rel_ns}}}Relationship"):
                rid = rel.get("Id", "")
                target = rel.get("Target", "")
                if "../media/" in target:
                    rid_map[rid] = "xl/media/" + target.split("../media/")[-1]
                elif "media/" in target:
                    rid_map[rid] = "xl/media/" + target.split("media/")[-1]
            drawing_rels[drawing_path] = rid_map

        img_counter = 1
        for drawing_path, rid_map in drawing_rels.items():
            if drawing_path not in names:
                continue
            tree = ET.parse(zf.open(drawing_path))
            root = tree.getroot()

            for anchor in root.findall(f"{{{ns['xdr']}}}twoCellAnchor") + \
                          root.findall(f"{{{ns['xdr']}}}oneCellAnchor"):
                # 행/열 위치 추출
                from_elem = anchor.find(f"{{{ns['xdr']}}}from")
                row = int(from_elem.find(f"{{{ns['xdr']}}}row").text) if from_elem is not None else -1
                col = int(from_elem.find(f"{{{ns['xdr']}}}col").text) if from_elem is not None else -1

                # rId 추출
                blip = anchor.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
                if blip is None:
                    continue
                rid = blip.get(f"{{{ns['r']}}}embed", "")
                media_path = rid_map.get(rid)
                if not media_path or media_path not in names:
                    continue

                ext = os.path.splitext(media_path)[1]
                filename = f"image{img_counter}{ext}"
                out_path = os.path.join(output_dir, filename)
                with open(out_path, "wb") as f:
                    f.write(zf.read(media_path))

                results.append({"filename": filename, "row": row, "col": col})
                img_counter += 1

        # ── Fallback: 드로잉 매핑 없는 경우 media 전체 추출 ──
        if not results:
            media_files = [n for n in names if n.startswith("xl/media/")]
            for media_path in media_files:
                ext = os.path.splitext(media_path)[1]
                filename = f"image{img_counter}{ext}"
                out_path = os.path.join(output_dir, filename)
                with open(out_path, "wb") as f:
                    f.write(zf.read(media_path))
                results.append({"filename": filename, "row": -1, "col": -1})
                img_counter += 1

    print(f"이미지 {len(results)}개 추출 완료 → {output_dir}")
    return results


# ──────────────────────────────────────────────────────────────────────────────
# 사용 예시 (직접 실행 시)
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("사용법: python excel_parser.py <파일.xlsx> [출력디렉토리]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_dir = sys.argv[2] if len(sys.argv) > 2 else "output"

    wb = read_workbook(xlsx_path)
    for sheet in wb.worksheets:
        print(f"\n=== 시트: {sheet.title} ===")
        header = detect_header_row(sheet)
        print(f"감지된 헤더 행: {header}")
        print(sheet_to_markdown(sheet, header_row=header))

        data = sheet_to_json(sheet, header_row=header)
        print(json.dumps(data, ensure_ascii=False, indent=2)[:500])

    images = extract_images(xlsx_path, out_dir)
    print(f"\n추출된 이미지: {images}")
