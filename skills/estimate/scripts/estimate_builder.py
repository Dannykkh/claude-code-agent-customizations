# requires: openpyxl
# pip install openpyxl

"""
개발 견적서 엑셀(.xlsx) 생성 헬퍼 라이브러리.

SKILL.md의 Phase 4(엑셀 파일 생성)에서 반복되는 openpyxl 패턴을 함수화합니다.
비용 그룹 추가 → 요약 시트 생성 → 수식 연결 → 저장의 흐름을 지원합니다.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# 스타일 상수
# ──────────────────────────────────────────────────────────────────────────────

_STYLE = {
    # 헤더: 진한 파란 배경 + 흰 글자
    "header_fill":    PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "header_font":    Font(color="FFFFFF", bold=True, size=11, name="맑은 고딕"),
    # 그룹 헤더: 연한 파란 배경
    "group_fill":     PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "group_font":     Font(bold=True, size=11, name="맑은 고딕"),
    # 소계: 연한 노란 배경
    "subtotal_fill":  PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "subtotal_font":  Font(bold=True, size=11, name="맑은 고딕"),
    # 합계: 진한 노란 배경
    "total_fill":     PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "total_font":     Font(bold=True, size=12, name="맑은 고딕"),
    # 기본 셀
    "normal_font":    Font(size=10, name="맑은 고딕"),
    # 테두리
    "thin_border":    Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    ),
}

# 금액 숫자 서식 (천 단위 콤마)
_NUMBER_FORMAT = '#,##0'


# ──────────────────────────────────────────────────────────────────────────────
# 워크북 생성
# ──────────────────────────────────────────────────────────────────────────────

def create_estimate_workbook(project_name: str) -> Workbook:
    """
    견적서 워크북을 생성하고 '견적서(Summary)' 시트에 컬럼 헤더를 설정합니다.

    생성 후에는 add_cost_group()으로 비용 항목을 추가하고,
    마지막에 add_summary_sheet()로 합계 시트를 붙인 뒤 저장하세요.

    Args:
        project_name: 프로젝트 이름 (시트 제목과 파일명에 사용)

    Returns:
        헤더가 설정된 Workbook 객체
    """
    wb = Workbook()

    # ── 기본 시트 제거 후 견적서 시트 생성 ──────────────────
    default_sheet = wb.active
    default_sheet.title = "견적서"

    ws = default_sheet
    ws.sheet_properties.tabColor = "1F4E79"

    # ── 제목 행 ──────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"{project_name} 개발 견적서"
    title_cell.font = Font(bold=True, size=14, name="맑은 고딕")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── 작성일 ────────────────────────────────────────────────
    ws["A2"] = f"작성일: {datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = Font(size=10, name="맑은 고딕", color="666666")
    ws.row_dimensions[2].height = 18

    # ── 컬럼 헤더 (3행) ───────────────────────────────────────
    headers = ["구분", "항목", "단위", "수량", "단가(원)", "금액(원)", "비고"]
    col_widths = [12, 30, 8, 8, 15, 15, 20]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = _STYLE["header_font"]
        cell.fill = _STYLE["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _STYLE["thin_border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 20

    # 행 포인터 (헬퍼 함수들이 공유): 4행부터 데이터 시작
    ws._estimate_row = 4
    wb._project_name = project_name
    wb._subtotal_refs: list[str] = []  # 소계 셀 참조 목록 (요약 시트용)

    return wb


# ──────────────────────────────────────────────────────────────────────────────
# 비용 그룹 추가
# ──────────────────────────────────────────────────────────────────────────────

def add_cost_group(ws: Worksheet, group_name: str, items: list[dict]) -> str:
    """
    견적서 시트에 비용 그룹과 항목들을 추가하고, 소계 행을 반환합니다.

    items 형식:
    [
        {"name": "기능A 개발", "unit": "MM", "qty": 0.5, "price": 8000000, "note": "보통"},
        {"name": "기능B 개발", "unit": "MM", "qty": 0.3, "price": 8000000},
    ]

    Args:
        ws:         create_estimate_workbook()으로 만든 '견적서' Worksheet
        group_name: 비용 그룹 이름. 예: "1. 개발비", "2. 인건비"
        items:      항목 목록. 각 항목은 name, unit, qty, price, note(선택) 포함.

    Returns:
        소계 셀 주소 (예: "F15"). add_summary_sheet()에서 합계 계산에 사용됩니다.
    """
    row = ws._estimate_row

    # ── 그룹 헤더 행 ──────────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws[f"A{row}"]
    cell.value = group_name
    cell.font = _STYLE["group_font"]
    cell.fill = _STYLE["group_fill"]
    cell.border = _STYLE["thin_border"]
    ws.row_dimensions[row].height = 18
    row += 1

    # ── 항목 행 ──────────────────────────────────────────────
    item_rows: list[int] = []
    for item in items:
        # A: 구분(빈칸), B: 항목명, C: 단위, D: 수량, E: 단가, F: 금액(수식), G: 비고
        ws[f"A{row}"].border = _STYLE["thin_border"]
        ws[f"B{row}"] = item.get("name", "")
        ws[f"C{row}"] = item.get("unit", "식")
        ws[f"D{row}"] = item.get("qty", 1)
        ws[f"E{row}"] = item.get("price", 0)
        ws[f"F{row}"] = f"=D{row}*E{row}"  # 수량 × 단가 수식
        ws[f"G{row}"] = item.get("note", "")

        # 스타일
        for col_letter in "ABCDEFG":
            c = ws[f"{col_letter}{row}"]
            c.font = _STYLE["normal_font"]
            c.border = _STYLE["thin_border"]
        ws[f"D{row}"].alignment = Alignment(horizontal="center")
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        ws[f"E{row}"].number_format = _NUMBER_FORMAT
        ws[f"F{row}"].number_format = _NUMBER_FORMAT

        item_rows.append(row)
        ws.row_dimensions[row].height = 16
        row += 1

    # ── 소계 행 ──────────────────────────────────────────────
    if item_rows:
        first, last = item_rows[0], item_rows[-1]
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = "소계"
        ws[f"A{row}"].font = _STYLE["subtotal_font"]
        ws[f"A{row}"].fill = _STYLE["subtotal_fill"]
        ws[f"A{row}"].alignment = Alignment(horizontal="right")
        ws[f"F{row}"] = f"=SUM(F{first}:F{last})"
        ws[f"F{row}"].number_format = _NUMBER_FORMAT
        ws[f"F{row}"].font = _STYLE["subtotal_font"]
        ws[f"F{row}"].fill = _STYLE["subtotal_fill"]
        ws[f"G{row}"].fill = _STYLE["subtotal_fill"]
        for col_letter in "ABCDEFG":
            ws[f"{col_letter}{row}"].border = _STYLE["thin_border"]
        ws.row_dimensions[row].height = 18

    subtotal_ref = f"F{row}"
    row += 2  # 그룹 사이 공백

    ws._estimate_row = row
    return subtotal_ref


# ──────────────────────────────────────────────────────────────────────────────
# 요약 시트 생성
# ──────────────────────────────────────────────────────────────────────────────

def add_summary_sheet(
    wb: Workbook,
    subtotal_refs: list[tuple[str, str]],
    buffer_pct: float = 10.0,
    vat_pct: float = 10.0,
) -> Worksheet:
    """
    각 비용 그룹의 소계를 모아 최종 합계를 보여주는 '요약' 시트를 추가합니다.

    예비비(버퍼)와 부가세(VAT) 수식도 자동으로 연결됩니다.

    Args:
        wb:             Workbook 객체
        subtotal_refs:  [(그룹명, 소계셀주소), ...] 목록.
                        add_cost_group() 반환값과 그룹명을 묶어서 전달합니다.
                        예: [("1. 개발비", "견적서!F15"), ("2. 인건비", "견적서!F22")]
        buffer_pct:     예비비 비율 (%). 기본값 10.
        vat_pct:        부가세 비율 (%). 기본값 10.

    Returns:
        생성된 '요약' Worksheet
    """
    ws = wb.create_sheet("요약")
    ws.sheet_properties.tabColor = "FFD966"

    # 제목
    ws.merge_cells("A1:C1")
    ws["A1"] = "견적 요약"
    ws["A1"].font = Font(bold=True, size=14, name="맑은 고딕")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 컬럼 헤더
    for col, (header, width) in enumerate(
        [("항목", 30), ("금액(원)", 18), ("비고", 20)], start=1
    ):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = _STYLE["header_font"]
        cell.fill = _STYLE["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = _STYLE["thin_border"]
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 18

    # 그룹 소계 행
    data_rows: list[int] = []
    for row_idx, (group_name, ref) in enumerate(subtotal_refs, start=3):
        ws[f"A{row_idx}"] = group_name
        ws[f"B{row_idx}"] = f"=견적서!{ref.split('!')[-1]}" if "!" not in ref else f"={ref}"
        ws[f"B{row_idx}"].number_format = _NUMBER_FORMAT
        ws[f"C{row_idx}"] = "소계"
        for col_letter in "ABC":
            ws[f"{col_letter}{row_idx}"].font = _STYLE["normal_font"]
            ws[f"{col_letter}{row_idx}"].border = _STYLE["thin_border"]
        ws.row_dimensions[row_idx].height = 16
        data_rows.append(row_idx)

    # 소계 합산 행
    subtotal_row = data_rows[-1] + 2 if data_rows else 5
    ws.merge_cells(f"A{subtotal_row}:A{subtotal_row}")
    ws[f"A{subtotal_row}"] = "소계"
    ws[f"B{subtotal_row}"] = f"=SUM(B3:B{data_rows[-1]})" if data_rows else 0
    ws[f"B{subtotal_row}"].number_format = _NUMBER_FORMAT
    _apply_subtotal_style(ws, subtotal_row)

    # 예비비 행
    buffer_row = subtotal_row + 1
    ws[f"A{buffer_row}"] = f"예비비 ({buffer_pct:.0f}%)"
    ws[f"B{buffer_row}"] = f"=B{subtotal_row}*{buffer_pct / 100}"
    ws[f"B{buffer_row}"].number_format = _NUMBER_FORMAT
    ws[f"C{buffer_row}"] = "리스크 버퍼"
    for col_letter in "ABC":
        ws[f"{col_letter}{buffer_row}"].font = _STYLE["normal_font"]
        ws[f"{col_letter}{buffer_row}"].border = _STYLE["thin_border"]
    ws.row_dimensions[buffer_row].height = 16

    # 부가세 행
    vat_row = buffer_row + 1
    ws[f"A{vat_row}"] = f"부가세 ({vat_pct:.0f}%)"
    ws[f"B{vat_row}"] = f"=(B{subtotal_row}+B{buffer_row})*{vat_pct / 100}"
    ws[f"B{vat_row}"].number_format = _NUMBER_FORMAT
    ws[f"C{vat_row}"] = "VAT"
    for col_letter in "ABC":
        ws[f"{col_letter}{vat_row}"].font = _STYLE["normal_font"]
        ws[f"{col_letter}{vat_row}"].border = _STYLE["thin_border"]
    ws.row_dimensions[vat_row].height = 16

    # 총 합계 행
    total_row = vat_row + 1
    ws[f"A{total_row}"] = "총 견적"
    ws[f"B{total_row}"] = f"=B{subtotal_row}+B{buffer_row}+B{vat_row}"
    ws[f"B{total_row}"].number_format = _NUMBER_FORMAT
    ws[f"C{total_row}"] = "VAT 포함"
    for col_letter in "ABC":
        cell = ws[f"{col_letter}{total_row}"]
        cell.font = _STYLE["total_font"]
        cell.fill = _STYLE["total_fill"]
        cell.border = _STYLE["thin_border"]
    ws.row_dimensions[total_row].height = 22

    return ws


def _apply_subtotal_style(ws: Worksheet, row: int) -> None:
    """소계 행에 스타일을 적용하는 내부 헬퍼."""
    for col_letter in "ABC":
        cell = ws[f"{col_letter}{row}"]
        cell.font = _STYLE["subtotal_font"]
        cell.fill = _STYLE["subtotal_fill"]
        cell.border = _STYLE["thin_border"]
    ws.row_dimensions[row].height = 18


# ──────────────────────────────────────────────────────────────────────────────
# 통화 형식
# ──────────────────────────────────────────────────────────────────────────────

def format_currency(value: int | float, currency: str = "KRW") -> str:
    """
    숫자를 통화 형식 문자열로 변환합니다.

    Args:
        value:    변환할 숫자
        currency: "KRW" (기본) | "USD" | "JPY"

    Returns:
        형식화된 문자열. 예: "8,000,000원", "$10,000", "¥1,200,000"
    """
    symbols = {"KRW": ("", "원"), "USD": ("$", ""), "JPY": ("¥", "")}
    prefix, suffix = symbols.get(currency.upper(), ("", ""))
    formatted = f"{int(value):,}"
    return f"{prefix}{formatted}{suffix}"


# ──────────────────────────────────────────────────────────────────────────────
# 합계 자동 계산
# ──────────────────────────────────────────────────────────────────────────────

def calculate_totals(ws: Worksheet, price_col: str = "E", amount_col: str = "F") -> dict:
    """
    견적서 시트에서 수량×단가를 재계산하여 합계를 반환합니다.

    openpyxl은 수식을 실행하지 않으므로, 이 함수는 Python 레벨에서
    D열(수량) × E열(단가)을 직접 계산하여 합산합니다.
    수식 셀이 이미 있는 경우에는 재계산 결과를 검증 용도로 사용하세요.

    Args:
        ws:          '견적서' Worksheet
        price_col:   단가 열 문자 (기본: "E")
        amount_col:  금액 열 문자 (기본: "F")

    Returns:
        {"total": 합계, "items": [(행번호, 수량, 단가, 금액), ...]}
    """
    items = []
    total = 0

    for row in ws.iter_rows(min_row=4, values_only=False):
        qty_cell = row[3]   # D열 (0-based: 3)
        price_cell = row[4]  # E열
        if qty_cell.value is None or price_cell.value is None:
            continue
        try:
            qty = float(qty_cell.value)
            price = float(price_cell.value)
        except (TypeError, ValueError):
            continue
        amount = qty * price
        total += amount
        items.append((qty_cell.row, qty, price, amount))

    return {"total": total, "items": items}


# ──────────────────────────────────────────────────────────────────────────────
# 사용 예시 (직접 실행 시)
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    wb = create_estimate_workbook("샘플 프로젝트")
    ws_main = wb["견적서"]

    # 비용 그룹 추가
    ref1 = add_cost_group(ws_main, "1. 개발비", [
        {"name": "프론트엔드 - 대시보드",    "unit": "MM", "qty": 0.5, "price": 8_000_000, "note": "보통"},
        {"name": "프론트엔드 - 회원관리",    "unit": "MM", "qty": 0.3, "price": 8_000_000, "note": "단순"},
        {"name": "백엔드 - REST API",         "unit": "MM", "qty": 1.0, "price": 8_000_000, "note": "복잡"},
        {"name": "백엔드 - DB 설계",          "unit": "MM", "qty": 0.3, "price": 8_000_000, "note": "단순"},
    ])

    ref2 = add_cost_group(ws_main, "2. 인건비", [
        {"name": "PM/기획",    "unit": "식", "qty": 1, "price": 2_000_000, "note": "개발비의 10%"},
        {"name": "QA/테스트",  "unit": "식", "qty": 1, "price": 3_000_000, "note": "개발비의 15%"},
    ])

    ref3 = add_cost_group(ws_main, "3. 인프라 비용(월)", [
        {"name": "AWS EC2 t3.medium", "unit": "월", "qty": 1, "price": 50_000},
        {"name": "RDS MySQL",          "unit": "월", "qty": 1, "price": 80_000},
    ])

    # 요약 시트
    add_summary_sheet(wb, [
        ("1. 개발비",       f"견적서!{ref1}"),
        ("2. 인건비",       f"견적서!{ref2}"),
        ("3. 인프라 비용",  f"견적서!{ref3}"),
    ])

    # 합계 검증
    totals = calculate_totals(ws_main)
    print(f"계산된 합계: {format_currency(totals['total'])}")

    # 저장
    out_path = f"estimate-샘플프로젝트-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(out_path)
    print(f"견적서 저장 완료: {os.path.abspath(out_path)}")
